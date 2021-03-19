import os
import openpyxl

from datetime import datetime
from openpyxl.styles import colors
from openpyxl.styles import Color, Font

from django.conf import settings
from django.contrib.auth import get_user_model
from django.db import transaction

from accounts.models import (
	CrawlingDate,
	SearchTagList, SearchTagResult,
	SearchAccountList, SearchAccountResult,
)

User = get_user_model()

def get_and_create_tags():
	wb = openpyxl.load_workbook(settings.XLSX_FILE_PATH)
	sheet = wb['Sheet1']
	rows, cols = sheet.max_row, sheet.max_column 

	user = User.objects.last()
	tag_list = []

	for i in range(1, int(rows)):
		tag = sheet.cell(int(i)+1, 1).value
		tag_list.append(SearchTagList(user=user, name=str(tag)))

	SearchTagList.objects.bulk_create(tag_list)


def get_and_create_accounts():
	wb = openpyxl.load_workbook(settings.XLSX_FILE_PATH)
	sheet = wb['Sheet2']
	rows, cols = sheet.max_row, sheet.max_column 

	user = User.objects.last()
	account_list = []

	for i in range(1, int(rows)):
		account = sheet.cell(int(i)+1, 1).value
		account_list.append(SearchAccountList(user=user, name=str(account)))

	SearchAccountList.objects.bulk_create(account_list)

@transaction.atomic
def save_tag_result():
	wb = openpyxl.load_workbook(settings.XLSX_FILE_PATH)
	sheet = wb['Sheet1']

	date_list = [
		datetime(2021, 2, 16),
		datetime(2021, 2, 17),
		datetime(2021, 2, 18),
		datetime(2021, 2, 22),
		datetime(2021, 2, 23),
		datetime(2021, 2, 24),
		datetime(2021, 2, 25),
		datetime(2021, 2, 26),
		datetime(2021, 3, 2),
		datetime(2021, 3, 3),
		datetime(2021, 3, 4),
		datetime(2021, 3, 5),
		datetime(2021, 3, 8),
		datetime(2021, 3, 9),
		datetime(2021, 3, 10),
		datetime(2021, 3, 11),
		datetime(2021, 3, 12),
		datetime(2021, 3, 15),
		datetime(2021, 3, 16),
		datetime(2021, 3, 17),
		datetime(2021, 3, 18),
		datetime(2021, 3, 19)
	]

	user = User.objects.last()
	tag_list = SearchTagList.objects.all()

	count_num = 2
	compare_num = 3

	for idx, date in enumerate(date_list):
		crawling_date = CrawlingDate.objects.create(
			user=user, crawling_type='tags', created_at=date
		)

		for idx, tag in enumerate(tag_list):
			search_tag = SearchTagList.objects.get(id=tag.id)
			count = sheet.cell((idx)+2, int(count_num)).value
			compare = sheet.cell((idx)+2, int(compare_num)).value
			
			SearchTagResult.objects.create(
				name=search_tag, count=int(count), compare=int(compare), 
				crawling_at=crawling_date, created_at=date
			)
		
		count_num = count_num + 2
		compare_num = compare_num + 2

@transaction.atomic
def save_account_result():
	wb = openpyxl.load_workbook(settings.XLSX_FILE_PATH)
	sheet = wb['Sheet2']

	date_list = [
		datetime(2021, 2, 13),
		datetime(2021, 2, 24),
		datetime(2021, 2, 25),		
		datetime(2021, 2, 26),		
		datetime(2021, 3, 2),		
		datetime(2021, 3, 3),		
		datetime(2021, 3, 4),		
		datetime(2021, 3, 5),		
		datetime(2021, 3, 8),		
		datetime(2021, 3, 9),		
		datetime(2021, 3, 10),		
		datetime(2021, 3, 11),		
		datetime(2021, 3, 12),		
		datetime(2021, 3, 15),		
		datetime(2021, 3, 16),		
		datetime(2021, 3, 17),
		datetime(2021, 3, 18),
		datetime(2021, 3, 19)
	]

	user = User.objects.last()
	account_list = SearchAccountList.objects.all()

	count_num = 2
	compare_num = 3

	for idx, date in enumerate(date_list):
		crawling_date = CrawlingDate.objects.create(
			user=user, crawling_type='accounts', created_at=date
		)

		for idx, account in enumerate(account_list):
			search_account = SearchAccountList.objects.get(id=account.id)
			count = sheet.cell((idx)+2, int(count_num)).value
			compare = sheet.cell((idx)+2, int(compare_num)).value
			
			SearchAccountResult.objects.create(
				name=search_account, count=int(count), compare=int(compare), 
				crawling_at=crawling_date, created_at=date
			)

		count_num = count_num + 2
		compare_num = compare_num + 2